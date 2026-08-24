from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

NAVY   = "1B2B5B"
GOLD   = "D4A843"
WHITE  = "FFFFFF"
LIGHT  = "F0F4FF"
GREEN  = "2E7D32"
GREEN_L= "E8F5E9"
RED    = "C62828"
RED_L  = "FFEBEE"
AMBER  = "E65100"
AMBER_L= "FFF3E0"
GRAY   = "455A64"
GRAY_L = "ECEFF1"
PURPLE = "4A148C"
PURPLE_L="F3E5F5"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right_align():
    return Alignment(horizontal="right", vertical="center")

# ─── HOJA 1: RESUMEN EJECUTIVO ───────────────────────────────────────────────
ws = wb.active
ws.title = "Resumen Ejecutivo"

# Anchos de columna
col_widths = [2, 30, 22, 22, 22, 18, 18, 2]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Altura de filas por defecto
ws.row_dimensions[1].height = 6

# ── ENCABEZADO ──
ws.merge_cells("B2:G2")
ws["B2"] = "RESUMEN EJECUTIVO — ESTADO REAL DEL BOOTCAMP"
ws["B2"].font = Font(bold=True, color=WHITE, size=16, name="Calibri")
ws["B2"].fill = fill(NAVY)
ws["B2"].alignment = center()
ws.row_dimensions[2].height = 30

ws.merge_cells("B3:G3")
ws["B3"] = "Sinergéticos / Synergy for Education  ·  Real Jun 2026 + Club Sinergético  ·  12 Mayo 2026  ·  Evento: Junio 5·6·7, 2026"
ws["B3"].font = font(color=GOLD, size=10, italic=True)
ws["B3"].fill = fill(NAVY)
ws["B3"].alignment = center()
ws.row_dimensions[3].height = 18

ws.row_dimensions[4].height = 8

# ── SECCIÓN: KPIs PRINCIPALES ──
ws.merge_cells("B5:G5")
ws["B5"] = "▸  ESTADO ACTUAL VS META — KPIs PRINCIPALES"
ws["B5"].font = Font(bold=True, color=WHITE, size=11, name="Calibri")
ws["B5"].fill = fill(NAVY)
ws["B5"].alignment = left()
ws.row_dimensions[5].height = 20

headers_kpi = ["MÉTRICA", "HOY — 12 Mayo 2026", "META", "% AVANCE", "SEMÁFORO"]
for col, h in enumerate(headers_kpi, 2):
    cell = ws.cell(row=6, column=col, value=h)
    cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    cell.fill = fill("2C3E6B")
    cell.alignment = center()
    cell.border = border_thin()
ws.row_dimensions[6].height = 18

kpis = [
    ("Leads totales",           "39,879",      "77,556",     "51.4%",  "⚠️ En proceso",         AMBER_L, AMBER),
    ("Registros confirmados",   "12,105",      "28,800",     "42.0%",  "🔴 Rezagado",            RED_L,   RED),
    ("Tasa de confirmación",    "30.35%",      "39.86%",     "76.1%",  "🔴 Crítico (−9.5 pp)",   RED_L,   RED),
    ("Inversión publicitaria",  "$220,000",    "$1,500,000", "14.7%",  "✅ Margen amplio",        GREEN_L, GREEN),
    ("CPL blended (pauta+Club)","$22",         "$19.34",     "—",      "⚠️ Encima de meta",      AMBER_L, AMBER),
    ("Blacks vendidos",         "170",         "278",        "61.2%",  "⚠️ 108 por cerrar",      AMBER_L, AMBER),
    ("Ingreso bruto generado",  "$4,111,620",  "$6,723,708", "61.2%",  "⚠️ En proceso",          AMBER_L, AMBER),
    ("Utilidad neta estimada",  "$1,841,048",  "$7,133,914", "25.8%",  "⚠️ Ver cascada",         AMBER_L, AMBER),
]

for r, (metrica, hoy, meta, pct, semaforo, bg, fc) in enumerate(kpis, 7):
    data = [metrica, hoy, meta, pct, semaforo]
    for c, val in enumerate(data, 2):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = fill(bg)
        cell.font = Font(bold=(c == 3), color=fc if c == 3 else "222222", size=10, name="Calibri")
        cell.alignment = center()
        cell.border = border_thin()
    ws.row_dimensions[r].height = 17

ws.row_dimensions[15].height = 8

# ── SECCIÓN: CASCADA DE COSTOS ──
ws.merge_cells("B16:G16")
ws["B16"] = "▸  CASCADA DE COSTOS — RESULTADO REAL A HOY (PAUTA + CLUB)"
ws["B16"].font = Font(bold=True, color=WHITE, size=11, name="Calibri")
ws["B16"].fill = fill(PURPLE)
ws["B16"].alignment = left()
ws.row_dimensions[16].height = 20

for col, h in zip([2, 3], ["CONCEPTO", "MONTO (MXN)"]):
    cell = ws.cell(row=17, column=col, value=h)
    cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    cell.fill = fill(NAVY)
    cell.alignment = center()
    cell.border = border_thin()
ws.merge_cells("C17:G17")
ws.row_dimensions[17].height = 17

cascada = [
    ("Ingreso bruto HOY (170 trans. × $24,186 prom.)",   "+$4,111,620",  LIGHT,   NAVY,   False),
    ("− Cascada de pagos (−40%: comis. 15% + banco 4% + IVA 16% + merma 5%)", "−$1,644,648", AMBER_L, AMBER, False),
    ("Ingreso neto variable (60%)",                      "$2,466,972",   "D1ECF1", "0C5460", True),
    ("− Twilio / WhatsApp API (~8,148 leads × $13)",     "−$105,924",    AMBER_L, AMBER,   False),
    ("− LLM / IA tokens (variable, estimado)",           "−$150,000",    AMBER_L, AMBER,   False),
    ("− Speakers del evento (costo fijo)",               "−$150,000",    AMBER_L, AMBER,   False),
    ("− Inversión de pauta ejecutada",                   "−$220,000",    AMBER_L, AMBER,   False),
    ("★  UTILIDAD NETA A HOY",                          "$1,841,048",   GREEN_L, GREEN,   True),
]

for r, (concepto, monto, bg, fc, bold) in enumerate(cascada, 18):
    ws.merge_cells(f"C{r}:G{r}")
    c1 = ws.cell(row=r, column=2, value=concepto)
    c1.fill = fill(bg)
    c1.font = Font(bold=bold, color=fc, size=10, name="Calibri")
    c1.alignment = left()
    c1.border = border_thin()

    c2 = ws.cell(row=r, column=3, value=monto)
    c2.fill = fill(bg)
    c2.font = Font(bold=bold, color=fc, size=10, name="Calibri")
    c2.alignment = right_align()
    c2.border = border_thin()
    ws.row_dimensions[r].height = 17

ws.row_dimensions[26].height = 8

# ── SECCIÓN: COMPARATIVO META vs HOY ──
ws.merge_cells("B27:G27")
ws["B27"] = "▸  COMPARATIVO COMPLETO — META vs REAL HOY"
ws["B27"].font = Font(bold=True, color=WHITE, size=11, name="Calibri")
ws["B27"].fill = fill("8B6914")
ws["B27"].alignment = left()
ws.row_dimensions[27].height = 20

comp_headers = ["MÉTRICA", "META", "HOY — 12 Mayo", "PENDIENTE", "% AVANCE", "SEMÁFORO"]
for c, h in enumerate(comp_headers, 2):
    cell = ws.cell(row=28, column=c, value=h)
    cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    cell.fill = fill(NAVY)
    cell.alignment = center()
    cell.border = border_thin()
ws.row_dimensions[28].height = 17

def section_header(ws, row, label):
    ws.merge_cells(f"B{row}:G{row}")
    ws[f"B{row}"] = label
    ws[f"B{row}"].font = Font(bold=True, color="AAB4D4", size=9, name="Calibri")
    ws[f"B{row}"].fill = fill(NAVY)
    ws[f"B{row}"].alignment = left()
    ws.row_dimensions[row].height = 14

comp_data = [
    ("ADQUISICIÓN", None),
    ("Leads totales",            "77,556",      "39,879",     "37,677 pend.",  "51.4%",  "⚠️ En proceso",    LIGHT,   "444444"),
    ("Registros confirmados",    "28,800",      "12,105",     "16,695 pend.",  "42.0%",  "🔴 Rezagado",      RED_L,   RED),
    ("Tasa de confirmación",     "39.86%",      "30.35%",     "−9.5 puntos",   "76.1%",  "🔴 Crítico",       RED_L,   RED),
    ("Inversión publicitaria",   "$1,500,000",  "$220,000",   "$1,280,000 rest.","14.7%","✅ Margen amplio", GREEN_L, GREEN),
    ("CPL blended",              "$19.34",      "$22.00",     "+$2.66 s/meta", "—",      "⚠️ Mejorar",       AMBER_L, AMBER),
    ("EVENTO EN VIVO", None),
    ("Pico máx. audiencia",      "13,036",      "6,703",      "6,333 pend.",   "51.4%",  "⚠️ Proporcional", LIGHT,   "444444"),
    ("Tasa de asistencia",       "16.81%",      "16.81%",     "—",             "100%",   "✅ En meta",       GREEN_L, GREEN),
    ("VENTAS BLACK", None),
    ("Blacks simples",           "119",         "73",         "46 pend.",      "61.3%",  "⚠️ En proceso",    LIGHT,   "444444"),
    ("Blacks dobles",            "81",          "49",         "32 pend.",      "60.5%",  "⚠️ En proceso",    LIGHT,   "444444"),
    ("Total Blacks",             "278",         "170",        "108 por cerrar","61.2%",  "⚠️ 108 x cerrar",  AMBER_L, AMBER),
    ("Conv. Blacks/Asistentes",  "2.13%",       "2.53%",      "—",             "—",      "✅ Por encima",    GREEN_L, GREEN),
    ("FINANCIERO", None),
    ("Ingreso bruto generado",   "$6,723,708",  "$4,111,620", "$2,612,088",    "61.2%",  "⚠️ En proceso",    LIGHT,   "444444"),
    ("Utilidad neta estimada",   "$7,133,914",  "$1,841,048", "$5,292,866",    "25.8%",  "⚠️ Ver cascada",   AMBER_L, AMBER),
]

row = 29
for item in comp_data:
    metrica = item[0]
    if item[1] is None:
        section_header(ws, row, f"  {metrica}")
        row += 1
        continue
    _, meta, hoy, pend, pct, semaforo, bg, fc = item
    vals = [metrica, meta, hoy, pend, pct, semaforo]
    for c, val in enumerate(vals, 2):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill(bg)
        cell.font = Font(bold=(c == 4), color=fc if c in [4, 7] else "222222", size=10, name="Calibri")
        cell.alignment = center() if c > 2 else left()
        cell.border = border_thin()
    ws.row_dimensions[row].height = 17
    row += 1

row += 1

# ── SECCIÓN: BRECHA ──
ws.merge_cells(f"B{row}:G{row}")
ws[f"B{row}"] = "▸  BRECHA — LO QUE FALTA PARA CERRAR LA META"
ws[f"B{row}"].font = Font(bold=True, color=WHITE, size=11, name="Calibri")
ws[f"B{row}"].fill = fill(RED)
ws[f"B{row}"].alignment = left()
ws.row_dimensions[row].height = 20
row += 1

brechas = [
    ("Blacks por cerrar",        "108",          "en los próximos 24 días"),
    ("Confirmados por lograr",   "16,695",       "leads por registrar"),
    ("Leads por conseguir",      "37,677",       "con $1.28M disponibles de pauta"),
    ("Utilidad por generar",     "$5,292,866",   "MXN pendientes"),
]
for label, val, sub in brechas:
    ws.merge_cells(f"B{row}:C{row}")
    c1 = ws[f"B{row}"]
    c1.value = label
    c1.font = Font(bold=True, color=NAVY, size=10, name="Calibri")
    c1.fill = fill("FFF8E1")
    c1.alignment = left()
    c1.border = border_thin()

    ws.merge_cells(f"D{row}:E{row}")
    c2 = ws[f"D{row}"]
    c2.value = val
    c2.font = Font(bold=True, color=RED, size=14, name="Calibri")
    c2.fill = fill("FFF8E1")
    c2.alignment = center()
    c2.border = border_thin()

    ws.merge_cells(f"F{row}:G{row}")
    c3 = ws[f"F{row}"]
    c3.value = sub
    c3.font = Font(italic=True, color="666666", size=9, name="Calibri")
    c3.fill = fill("FFF8E1")
    c3.alignment = left()
    c3.border = border_thin()
    ws.row_dimensions[row].height = 18
    row += 1

row += 1

# ── SECCIÓN: PROYECCIONES ──
ws.merge_cells(f"B{row}:G{row}")
ws[f"B{row}"] = "▸  PROYECCIÓN — ¿ES ALCANZABLE LA META?"
ws[f"B{row}"].font = Font(bold=True, color=WHITE, size=11, name="Calibri")
ws[f"B{row}"].fill = fill(GREEN)
ws[f"B{row}"].alignment = left()
ws.row_dimensions[row].height = 20
row += 1

proy_headers = ["ESCENARIO", "SUPUESTO", "BLACKS TOTALES", "BRUTO EST.", "VEREDICTO"]
for c, h in enumerate(proy_headers, 2):
    cell = ws.cell(row=row, column=c, value=h)
    cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    cell.fill = fill(NAVY)
    cell.alignment = center()
    cell.border = border_thin()
ws.row_dimensions[row].height = 17
row += 1

proyecciones = [
    ("🔴 Pesimista — CPL $27",
     "$1.28M → 47,407 leads · Club 22K al 30% conv.",
     "~240", "$7.3M", "Meta en riesgo", RED_L, RED),
    ("🟡 Base — Club al 40%",
     "$1.28M a $27 CPL · Club 22K al 40% conf. · asist. 16.81%",
     "~278", "$8.6M", "Meta alcanzable", AMBER_L, AMBER),
    ("🎯 Óptimo — CPL baja a $20",
     "$1.28M → 64K leads · Club 40% · conv. evento 2.53%",
     "~310", "$9.4M", "Meta superada", GREEN_L, GREEN),
]
for escenario, supuesto, blacks, bruto, veredicto, bg, fc in proyecciones:
    vals = [escenario, supuesto, blacks, bruto, veredicto]
    for c, val in enumerate(vals, 2):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill(bg)
        cell.font = Font(bold=(c == 7), color=fc if c == 7 else "222222", size=10, name="Calibri")
        cell.alignment = center() if c > 2 else left()
        cell.border = border_thin()
    ws.row_dimensions[row].height = 17
    row += 1

row += 1

# ── SUPUESTOS ──
ws.merge_cells(f"B{row}:G{row}")
ws[f"B{row}"] = "▸  SUPUESTOS DEL MODELO"
ws[f"B{row}"].font = Font(bold=True, color=WHITE, size=11, name="Calibri")
ws[f"B{row}"].fill = fill(GRAY)
ws[f"B{row}"].alignment = left()
ws.row_dimensions[row].height = 20
row += 1

supuestos = [
    ("Cascada de pagos",      "−40% del ingreso bruto (comis. 15% + banco 4% + IVA 16% + merma 5%)"),
    ("Ticket promedio",       "$24,186 MXN / transacción (ingreso bruto ÷ transacciones)"),
    ("Twilio / WhatsApp",     "$13 MXN por lead — solo pauta (~8,148 leads: $220K ÷ $27 CPL)"),
    ("LLM / IA tokens",       "$150,000 estimado — variable según interacciones"),
    ("Speakers del evento",   "$150,000 MXN fijo"),
    ("Club Sinergético",      "22,000 leads a $0 CPL — leads orgánicos sin costo de pauta"),
    ("Fuente de datos",       "Dashboard Columna I 'HOY (REAL JUN 2026 + CLUB SINERGÉTICO)' · 12 mayo 2026"),
]
for key, val in supuestos:
    ws.merge_cells(f"B{row}:C{row}")
    c1 = ws[f"B{row}"]
    c1.value = key
    c1.font = Font(bold=True, color=NAVY, size=9, name="Calibri")
    c1.fill = fill(GRAY_L)
    c1.alignment = left()
    c1.border = border_thin()

    ws.merge_cells(f"D{row}:G{row}")
    c2 = ws[f"D{row}"]
    c2.value = val
    c2.font = font(size=9)
    c2.fill = fill(GRAY_L)
    c2.alignment = left()
    c2.border = border_thin()
    ws.row_dimensions[row].height = 16
    row += 1

# Footer
row += 1
ws.merge_cells(f"B{row}:G{row}")
ws[f"B{row}"] = "Sinergéticos / Synergy for Education  ·  Bootcamp Digital Junio 5–7, 2026  ·  Resumen ejecutivo — 12 mayo 2026  ·  Fuente: Dashboard Columna I"
ws[f"B{row}"].font = Font(italic=True, color=WHITE, size=8, name="Calibri")
ws[f"B{row}"].fill = fill(NAVY)
ws[f"B{row}"].alignment = center()
ws.row_dimensions[row].height = 16

# Freeze panes
ws.freeze_panes = "B7"

output_path = "/Users/pablomunoz/Bootcamp/Resumen_Ejecutivo_HOY_May2026.xlsx"
wb.save(output_path)
print(f"Archivo guardado: {output_path}")
