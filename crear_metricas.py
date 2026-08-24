import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule

# ─── COLORES DE MARCA ────────────────────────────────────────────────
NEGRO       = "FF1A1A1A"
ROJO        = "FFD63B2F"
NARANJA     = "FFE8640C"
BLANCO      = "FFFFFFFF"
GRIS_OSC    = "FF2D2D2D"
GRIS_MED    = "FF444444"
GRIS_CLARO  = "FFF5F5F5"
VERDE       = "FF2ECC71"
AMARILLO    = "FFF1C40F"
AZUL        = "FF2980B9"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=BLANCO, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def border_thin():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom():
    s = Side(style="medium", color="FF888888")
    return Border(bottom=s)

def style_header(ws, row, col, value, bg=ROJO, fg=BLANCO, size=12, bold=True, span=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=size)
    cell.alignment = center()
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return cell

def style_label(ws, row, col, value, bg=GRIS_OSC, fg=BLANCO, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=10)
    cell.alignment = left()
    cell.border = border_thin()
    return cell

def style_value(ws, row, col, value, bg=GRIS_CLARO, fg=NEGRO, bold=False, fmt=None, center_align=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=10)
    cell.alignment = center() if center_align else left()
    cell.border = border_thin()
    if fmt:
        cell.number_format = fmt
    return cell


wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════
# HOJA 1 — DASHBOARD
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📊 Dashboard"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 32
ws1.column_dimensions["C"].width = 20
ws1.column_dimensions["D"].width = 20
ws1.column_dimensions["E"].width = 20
ws1.column_dimensions["F"].width = 20
ws1.column_dimensions["G"].width = 3
ws1.row_dimensions[1].height = 10

# Título principal
ws1.merge_cells("B2:F2")
c = ws1["B2"]
c.value = "BOOTCAMP — MÉTRICAS Y VENTAS"
c.fill = fill(NEGRO)
c.font = Font(bold=True, color=ROJO[2:], size=20, name="Calibri")
c.alignment = center()
ws1.row_dimensions[2].height = 45

ws1.merge_cells("B3:F3")
c = ws1["B3"]
c.value = "Sinergéticos / Synergy for Education"
c.fill = fill(NEGRO)
c.font = font(bold=False, color=BLANCO[2:], size=11, italic=True)
c.alignment = center()
ws1.row_dimensions[3].height = 22

ws1.row_dimensions[4].height = 10

# Encabezados de comparativo
style_header(ws1, 5, 2, "MÉTRICA", bg=NEGRO, size=10)
style_header(ws1, 5, 3, "BOOTCAMP ENE 2025", bg=ROJO, size=10)
style_header(ws1, 5, 4, "META JUN 2026", bg=NARANJA, size=10)
style_header(ws1, 5, 5, "REAL JUN 2026", bg=GRIS_MED, size=10)
style_header(ws1, 5, 6, "VS META", bg=GRIS_OSC, size=10)
ws1.row_dimensions[5].height = 30

# ── Sección: ADQUISICIÓN ──
ws1.merge_cells("B6:F6")
c = ws1["B6"]
c.value = "▸  ADQUISICIÓN DE LEADS"
c.fill = fill(ROJO)
c.font = font(bold=True, color=BLANCO[2:], size=10)
c.alignment = left()
ws1.row_dimensions[6].height = 22

data_dash = [
    # label,              ene2025,     meta_jun,   fmt
    ("Leads totales",       58530,      70000,     '#,##0'),
    ("Registros confirmados", 23329,    28000,     '#,##0'),
    ("Tasa de confirmación", 0.3986,    0.40,      '0.00%'),
    ("Inversión publicitaria (MXN)", None, None,   '$#,##0'),
    ("Costo por lead (CPL)",  None,     None,      '$#,##0.00'),
]

row = 7
for label, v_ene, v_meta, fmt in data_dash:
    style_label(ws1, row, 2, label)
    style_value(ws1, row, 3, v_ene, fmt=fmt, center_align=True)
    style_value(ws1, row, 4, v_meta, bg="FFFEF9EF", fmt=fmt, center_align=True)
    style_value(ws1, row, 5, None, bg="FFEEEEEE", fmt=fmt, center_align=True)
    # Fórmula VS META
    f_cell = ws1.cell(row=row, column=6)
    f_cell.value = f'=IF(E{row}="","—",E{row}-D{row})'
    f_cell.fill = fill("FFEEEEEE")
    f_cell.font = font(color=NEGRO[2:], size=10)
    f_cell.alignment = center()
    f_cell.border = border_thin()
    f_cell.number_format = fmt if fmt else 'General'
    ws1.row_dimensions[row].height = 20
    row += 1

# ── Sección: EVENTO EN VIVO ──
ws1.merge_cells(f"B{row}:F{row}")
c = ws1.cell(row=row, column=2, value="▸  EVENTO EN VIVO")
c.fill = fill(NARANJA)
c.font = font(bold=True, color=BLANCO[2:], size=10)
c.alignment = left()
ws1.row_dimensions[row].height = 22
row += 1

data_evento = [
    ("Visualizaciones totales",   45000,   55000,  '#,##0'),
    ("Pico máximo de audiencia",   9099,   11000,  '#,##0'),
    ("Tasa de asistencia",        0.1555,   0.18,  '0.00%'),
    ("Interesados capturados Día 1",  239,   300,  '#,##0'),
    ("Interesados capturados Día 2",  230,   290,  '#,##0'),
    ("Retención Día 1 → Día 2",   None,    None,   '0.00%'),
    ("Retención Día 2 → Día 3",   None,    None,   '0.00%'),
]
for label, v_ene, v_meta, fmt in data_evento:
    style_label(ws1, row, 2, label)
    style_value(ws1, row, 3, v_ene, fmt=fmt, center_align=True)
    style_value(ws1, row, 4, v_meta, bg="FFFEF9EF", fmt=fmt, center_align=True)
    style_value(ws1, row, 5, None, bg="FFEEEEEE", fmt=fmt, center_align=True)
    f_cell = ws1.cell(row=row, column=6)
    f_cell.value = f'=IF(E{row}="","—",E{row}-D{row})'
    f_cell.fill = fill("FFEEEEEE")
    f_cell.font = font(color=NEGRO[2:], size=10)
    f_cell.alignment = center()
    f_cell.border = border_thin()
    f_cell.number_format = fmt
    ws1.row_dimensions[row].height = 20
    row += 1

# ── Sección: VENTAS BLACK ──
ws1.merge_cells(f"B{row}:F{row}")
c = ws1.cell(row=row, column=2, value="▸  VENTAS SYNERGY BLACK ACCESS")
c.fill = fill(NEGRO)
c.font = font(bold=True, color=ROJO[2:], size=10)
c.alignment = left()
ws1.row_dimensions[row].height = 22
row += 1

data_ventas = [
    ("Blacks simples (completos)",    78,    90,   '#,##0'),
    ("Blacks dobles (completos)",    110,   130,   '#,##0'),
    ("Apartados simples",              8,    15,   '#,##0'),
    ("Apartados dobles",              14,    20,   '#,##0'),
    ("Total transacciones Black",    210,   250,   '#,##0'),
    ("Total accesos vendidos",       298,   380,   '#,##0'),
    ("Tasa conversión (Blacks/Leads)", 0.00359, 0.0036, '0.000%'),
    ("Tasa conversión (Blacks/Confirmados)", 0.009, 0.009, '0.00%'),
    ("Tasa conversión (Blacks/Asistentes)", 0.00467, 0.005, '0.00%'),
    ("Ingreso neto completos (MXN)", 4800929.02, 6000000, '$#,##0.00'),
    ("Ingreso neto apartados (MXN)",  131677.73,  200000, '$#,##0.00'),
    ("Total ingreso neto (MXN)",    4932606.75, 6200000, '$#,##0.00'),
    ("Por cobrar — apartados (MXN)",  735302.26,  None,  '$#,##0.00'),
    ("ROI de campaña",                None,       None,  '0.00%'),
    ("Costo por Black vendido (MXN)", None,       None,  '$#,##0'),
]
for label, v_ene, v_meta, fmt in data_ventas:
    style_label(ws1, row, 2, label)
    style_value(ws1, row, 3, v_ene, fmt=fmt, center_align=True)
    style_value(ws1, row, 4, v_meta, bg="FFFEF9EF", fmt=fmt, center_align=True)
    style_value(ws1, row, 5, None, bg="FFEEEEEE", fmt=fmt, center_align=True)
    f_cell = ws1.cell(row=row, column=6)
    f_cell.value = f'=IF(E{row}="","—",E{row}-D{row})'
    f_cell.fill = fill("FFEEEEEE")
    f_cell.font = font(color=NEGRO[2:], size=10)
    f_cell.alignment = center()
    f_cell.border = border_thin()
    f_cell.number_format = fmt
    ws1.row_dimensions[row].height = 20
    row += 1

# ── Nota al pie ──
row += 1
ws1.merge_cells(f"B{row}:F{row}")
c = ws1.cell(row=row, column=2)
c.value = "💡 Las metas Jun 2026 se calcularon con un crecimiento estimado del 20% sobre Ene 2025. Ajústalas según tu estrategia."
c.fill = fill("FFFFF8E1")
c.font = Font(italic=True, color="FF795548", size=9, name="Calibri")
c.alignment = left()
ws1.row_dimensions[row].height = 30


# ════════════════════════════════════════════════════════════════════
# HOJA 2 — LEADS & CANALES
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🎯 Leads & Canales")
ws2.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHI", [3,28,16,16,16,16,16,16,3]):
    ws2.column_dimensions[get_column_letter(ord(col)-64)].width = w

ws2.row_dimensions[1].height = 10
ws2.merge_cells("B2:H2")
c = ws2["B2"]
c.value = "LEADS & CANALES DE ADQUISICIÓN — BOOTCAMP JUN 2026"
c.fill = fill(NEGRO)
c.font = Font(bold=True, color=ROJO[2:], size=16, name="Calibri")
c.alignment = center()
ws2.row_dimensions[2].height = 40

# Sub-headers
headers = ["CANAL", "LEADS ENE 2025", "META JUN 2026", "REAL JUN 2026",
           "COSTO TOTAL ($)", "CPL ($)", "% DEL TOTAL"]
for i, h in enumerate(headers, 2):
    style_header(ws2, 4, i, h, bg=ROJO if i==2 else NARANJA if i==3 else NEGRO, size=10)
ws2.row_dimensions[4].height = 28

canales = [
    ("Meta Ads (Facebook/Instagram)", 0, 0),
    ("TikTok Ads", 0, 0),
    ("Orgánico Instagram", 0, 0),
    ("Orgánico Facebook", 0, 0),
    ("Orgánico TikTok", 0, 0),
    ("Email Marketing", 0, 0),
    ("WhatsApp / Referidos", 0, 0),
    ("YouTube / Podcast", 0, 0),
    ("Otros", 0, 0),
]

row = 5
for canal, v_ene, v_meta in canales:
    bg = GRIS_CLARO if row % 2 == 0 else BLANCO
    style_label(ws2, row, 2, canal, bg=GRIS_OSC)
    style_value(ws2, row, 3, v_ene if v_ene else None, bg=bg, fmt='#,##0', center_align=True)
    style_value(ws2, row, 4, v_meta if v_meta else None, bg="FFFEF9EF", fmt='#,##0', center_align=True)
    style_value(ws2, row, 5, None, bg="FFEEEEEE", fmt='#,##0', center_align=True)
    style_value(ws2, row, 6, None, bg="FFEEEEEE", fmt='$#,##0', center_align=True)
    style_value(ws2, row, 7, None, bg="FFEEEEEE", fmt='$#,##0.00', center_align=True)
    # % del total
    pct_cell = ws2.cell(row=row, column=8)
    pct_cell.value = f"=IF(SUM(E5:E13)=0,\"\",E{row}/SUM($E$5:$E$13))"
    pct_cell.fill = fill("FFEEEEEE")
    pct_cell.font = font(color=NEGRO[2:], size=10)
    pct_cell.alignment = center()
    pct_cell.border = border_thin()
    pct_cell.number_format = '0.0%'
    ws2.row_dimensions[row].height = 20
    row += 1

# Fila total
style_header(ws2, row, 2, "TOTAL", bg=NEGRO, size=10)
for col_idx in range(3, 9):
    c = ws2.cell(row=row, column=col_idx)
    col_letter = get_column_letter(col_idx)
    c.value = f"=SUM({col_letter}5:{col_letter}{row-1})" if col_idx != 8 else "100%"
    c.fill = fill(NEGRO)
    c.font = font(bold=True, color=BLANCO[2:], size=10)
    c.alignment = center()
    c.border = border_thin()
    fmts = {3:'#,##0', 4:'#,##0', 5:'#,##0', 6:'$#,##0', 7:'$#,##0.00', 8:'0.0%'}
    c.number_format = fmts.get(col_idx, 'General')
ws2.row_dimensions[row].height = 24
row += 2

# Sección: tracking diario de registros
ws2.merge_cells(f"B{row}:H{row}")
c = ws2.cell(row=row, column=2)
c.value = "REGISTRO ACUMULADO POR DÍA (Pre-evento)"
c.fill = fill(NARANJA)
c.font = font(bold=True, size=11)
c.alignment = center()
ws2.row_dimensions[row].height = 26
row += 1

day_headers = ["FECHA", "LEADS DEL DÍA", "LEADS ACUMULADOS", "CONFIRMADOS DEL DÍA",
               "CONFIRMADOS ACUM.", "TASA CONFIRM. DIARIA", ""]
for i, h in enumerate(day_headers, 2):
    style_header(ws2, row, i, h, bg=GRIS_OSC, size=9)
ws2.row_dimensions[row].height = 26
row += 1

for semana in range(1, 29):
    bg = GRIS_CLARO if semana % 2 == 0 else BLANCO
    style_value(ws2, row, 2, f"Día {semana}", bg=GRIS_OSC, fg=BLANCO, center_align=True)
    for col in range(3, 8):
        c_val = ws2.cell(row=row, column=col)
        c_val.fill = fill(bg)
        c_val.font = font(color=NEGRO[2:], size=10)
        c_val.alignment = center()
        c_val.border = border_thin()
        fmts = {3:'#,##0', 4:'#,##0', 5:'#,##0', 6:'#,##0', 7:'0.00%'}
        c_val.number_format = fmts.get(col, 'General')
    ws2.row_dimensions[row].height = 18
    row += 1


# ════════════════════════════════════════════════════════════════════
# HOJA 3 — EVENTO EN VIVO
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🔴 Evento en Vivo")
ws3.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHI", [3,30,18,18,18,18,18,3,3]):
    ws3.column_dimensions[get_column_letter(ord(col)-64)].width = w

ws3.row_dimensions[1].height = 10
ws3.merge_cells("B2:G2")
c = ws3["B2"]
c.value = "MÉTRICAS EVENTO EN VIVO — 5, 6 Y 7 DE JUNIO 2026"
c.fill = fill(NEGRO)
c.font = Font(bold=True, color=ROJO[2:], size=16, name="Calibri")
c.alignment = center()
ws3.row_dimensions[2].height = 40

dias = [
    ("DÍA 1 — Jueves 5 Jun", "FF8B0000"),
    ("DÍA 2 — Viernes 6 Jun", "FFB71C1C"),
    ("DÍA 3 — Sábado 7 Jun", "FFD32F2F"),
]

metricas_dia = [
    ("Visualizaciones totales", '#,##0', [45000, None, None]),
    ("Pico máximo de audiencia", '#,##0', [9099, None, None]),
    ("Visualizaciones únicas estimadas", '#,##0', [None, None, None]),
    ("Tasa de asistencia (viz/confirmados)", '0.00%', [None, None, None]),
    ("Retención vs día anterior", '0.00%', [None, None, None]),
    ("Interesados capturados", '#,##0', [239, 230, None]),
    ("Preguntas / interacciones", '#,##0', [None, None, None]),
    ("Comentarios en vivo", '#,##0', [None, None, None]),
]

row = 4
style_header(ws3, row, 2, "MÉTRICA", bg=NEGRO, size=10)
for i, (dia, color) in enumerate(dias):
    style_header(ws3, row, 3+i, dia, bg=color[2:] if len(color) > 2 else color, size=10)
    ws3.cell(row=row, column=3+i).fill = fill(color)
style_header(ws3, row, 6, "TOTAL / PROMEDIO", bg=NARANJA, size=10)
ws3.row_dimensions[row].height = 30
row += 1

for label, fmt, vals in metricas_dia:
    style_label(ws3, row, 2, label)
    for i, v in enumerate(vals):
        style_value(ws3, row, 3+i, v, fmt=fmt, center_align=True)
    # Total/prom
    total_cell = ws3.cell(row=row, column=6)
    if '%' in fmt:
        total_cell.value = f"=AVERAGE(C{row}:E{row})"
    else:
        total_cell.value = f"=SUM(C{row}:E{row})"
    total_cell.fill = fill("FFFEF9EF")
    total_cell.font = font(bold=True, color=NEGRO[2:], size=10)
    total_cell.alignment = center()
    total_cell.border = border_thin()
    total_cell.number_format = fmt
    ws3.row_dimensions[row].height = 20
    row += 1

row += 1
# Sección oferta
ws3.merge_cells(f"B{row}:F{row}")
c = ws3.cell(row=row, column=2)
c.value = "MOMENTO DE LA OFERTA — SYNERGY BLACK ACCESS"
c.fill = fill(NEGRO)
c.font = font(bold=True, color=ROJO[2:], size=11)
c.alignment = center()
ws3.row_dimensions[row].height = 26
row += 1

oferta_headers = ["MOMENTO", "DÍA", "HORA", "TIPO OFERTA", "BLACKS VENDIDOS", "INGRESO ($)"]
for i, h in enumerate(oferta_headers, 2):
    style_header(ws3, row, i, h, bg=GRIS_OSC, size=9)
ws3.row_dimensions[row].height = 24
row += 1

momentos = [
    ("Apertura early bird", "Día 1", "09:00", "Black Simple / Doble"),
    ("Cierre Día 1", "Día 1", "19:00", "Black Simple / Doble"),
    ("Apertura Día 2", "Día 2", "09:00", "Black Simple / Doble"),
    ("Cierre Día 2 + urgencia", "Día 2", "19:00", "Black Simple / Doble"),
    ("Última llamada Día 3", "Día 3", "12:00", "Black Simple / Doble"),
    ("Cierre final", "Día 3", "19:00", "Black Simple / Doble"),
]
for momento, dia, hora, tipo in momentos:
    bg = GRIS_CLARO if row % 2 == 0 else BLANCO
    style_value(ws3, row, 2, momento, bg=GRIS_OSC, fg=BLANCO)
    style_value(ws3, row, 3, dia, bg=bg, center_align=True)
    style_value(ws3, row, 4, hora, bg=bg, center_align=True)
    style_value(ws3, row, 5, tipo, bg=bg, center_align=True)
    style_value(ws3, row, 6, None, bg="FFEEEEEE", fmt='#,##0', center_align=True)
    style_value(ws3, row, 7, None, bg="FFEEEEEE", fmt='$#,##0', center_align=True)
    ws3.row_dimensions[row].height = 20
    row += 1


# ════════════════════════════════════════════════════════════════════
# HOJA 4 — VENTAS BLACK
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("⚫ Ventas Black")
ws4.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJ", [3,28,16,16,16,16,16,16,16,3]):
    ws4.column_dimensions[get_column_letter(ord(col)-64)].width = w

ws4.row_dimensions[1].height = 10
ws4.merge_cells("B2:I2")
c = ws4["B2"]
c.value = "VENTAS SYNERGY BLACK ACCESS — TRACKER COMPLETO"
c.fill = fill(NEGRO)
c.font = Font(bold=True, color=ROJO[2:], size=16, name="Calibri")
c.alignment = center()
ws4.row_dimensions[2].height = 40

# Precios de referencia
ws4.row_dimensions[4].height = 20
ws4.merge_cells("B4:D4")
ws4["B4"].value = "💰 PRECIOS DE REFERENCIA"
ws4["B4"].fill = fill(GRIS_OSC)
ws4["B4"].font = font(bold=True, size=10)
ws4["B4"].alignment = center()

precios_ref = [
    ("Black Simple", 36997, "1 acceso"),
    ("Black Doble", 55497, "2 accesos — AHORRO $18,497"),
]
row = 5
for nombre, precio, nota in precios_ref:
    style_value(ws4, row, 2, nombre, bg=GRIS_CLARO, bold=True)
    style_value(ws4, row, 3, precio, bg=GRIS_CLARO, fmt='$#,##0', center_align=True)
    style_value(ws4, row, 4, nota, bg=GRIS_CLARO, center_align=True)
    ws4.row_dimensions[row].height = 20
    row += 1

row += 1
# Tabla principal comparativa
headers4 = ["TIPO", "ENE 2025\nCompletos", "ENE 2025\nApartados", "ENE 2025\nTotal",
            "META JUN 2026", "REAL JUN 2026\nCompletos", "REAL JUN 2026\nApartados",
            "REAL JUN 2026\nTotal"]
for i, h in enumerate(headers4, 2):
    style_header(ws4, row, i, h, bg=ROJO if i <= 5 else NARANJA, size=9)
ws4.row_dimensions[row].height = 36
row += 1

tipos_black = [
    ("Black Simple", 78, 8, 86, 100, None, None),
    ("Black Doble (2 accesos)", 110, 14, 124, 130, None, None),
    ("TOTAL TRANSACCIONES", 188, 22, 210, 230, None, None),
    ("TOTAL ACCESOS (doble×2)", 298, None, 298, 380, None, None),
]
for t_data in tipos_black:
    label = t_data[0]
    vals = t_data[1:]
    is_total = "TOTAL" in label
    bg = NEGRO if is_total else GRIS_CLARO
    fg = BLANCO if is_total else NEGRO
    bold = is_total
    style_label(ws4, row, 2, label, bg=bg, fg=fg, bold=bold)
    for i, v in enumerate(vals):
        c = ws4.cell(row=row, column=3+i)
        c.value = v
        c.fill = fill(NEGRO if is_total else ("FFEEEEEE" if i >= 4 else GRIS_CLARO))
        c.font = font(bold=bold, color=(BLANCO if is_total else NEGRO)[2:], size=10)
        c.alignment = center()
        c.border = border_thin()
        c.number_format = '#,##0'
    ws4.row_dimensions[row].height = 22
    row += 1

row += 1
# Sección ingresos
ws4.merge_cells(f"B{row}:I{row}")
c = ws4.cell(row=row, column=2)
c.value = "INGRESOS"
c.fill = fill(NEGRO)
c.font = font(bold=True, color=ROJO[2:], size=11)
c.alignment = center()
ws4.row_dimensions[row].height = 26
row += 1

ingresos_headers = ["CONCEPTO", "ENE 2025", "META JUN 2026", "REAL JUN 2026", "DIFERENCIA VS META", "", "", ""]
for i, h in enumerate(ingresos_headers, 2):
    if h:
        style_header(ws4, row, i, h, bg=GRIS_OSC, size=9)
ws4.row_dimensions[row].height = 24
row += 1

ingresos = [
    ("Ingreso neto completos", 4800929.02, 6000000),
    ("Ingreso neto apartados", 131677.73, 200000),
    ("Total ingreso neto", 4932606.75, 6200000),
    ("Por cobrar (apartados pendientes)", 735302.26, None),
    ("Potencial recuperación (80%)", 919127.82, None),
    ("Ingreso proyectado total", 5667908.57, None),
]
for label, v_ene, v_meta in ingresos:
    is_total = "Total" in label or "Potencial" in label or "proyectado" in label
    bg = GRIS_OSC if is_total else GRIS_CLARO
    fg = BLANCO if is_total else NEGRO
    style_label(ws4, row, 2, label, bg=bg, fg=fg, bold=is_total)
    style_value(ws4, row, 3, v_ene, bg=bg if is_total else GRIS_CLARO,
                fg=fg, fmt='$#,##0.00', center_align=True, bold=is_total)
    style_value(ws4, row, 4, v_meta, bg="FFFEF9EF", fmt='$#,##0.00', center_align=True)
    style_value(ws4, row, 5, None, bg="FFEEEEEE", fmt='$#,##0.00', center_align=True)
    diff_cell = ws4.cell(row=row, column=6)
    diff_cell.value = f'=IF(E{row}="","—",E{row}-D{row})'
    diff_cell.fill = fill("FFEEEEEE")
    diff_cell.font = font(color=NEGRO[2:], size=10)
    diff_cell.alignment = center()
    diff_cell.border = border_thin()
    diff_cell.number_format = '$#,##0.00'
    ws4.row_dimensions[row].height = 20
    row += 1

row += 1
# KPIs de conversión
ws4.merge_cells(f"B{row}:I{row}")
c = ws4.cell(row=row, column=2)
c.value = "KPIs DE CONVERSIÓN"
c.fill = fill(NARANJA)
c.font = font(bold=True, size=11)
c.alignment = center()
ws4.row_dimensions[row].height = 26
row += 1

kpis_conv = [
    ("Tasa conversión Blacks / Leads", "0.36%", "0.36%", '0.000%'),
    ("Tasa conversión Blacks / Confirmados", "0.90%", "0.89%", '0.00%'),
    ("Tasa conversión Blacks / Asistentes", "0.47%", "0.45%", '0.00%'),
    ("Ticket promedio por transacción", "$23,489", "$24,800", '$#,##0'),
    ("Costo por Black vendido (MXN)", "—", "—", '$#,##0'),
    ("ROI de la campaña", "—", "—", '0.00%'),
]
for label, v_ene, v_meta, fmt in kpis_conv:
    style_label(ws4, row, 2, label)
    style_value(ws4, row, 3, v_ene, center_align=True)
    style_value(ws4, row, 4, v_meta, bg="FFFEF9EF", center_align=True)
    style_value(ws4, row, 5, None, bg="FFEEEEEE", center_align=True)
    style_value(ws4, row, 6, None, bg="FFEEEEEE", center_align=True)
    ws4.row_dimensions[row].height = 20
    row += 1


# ════════════════════════════════════════════════════════════════════
# HOJA 5 — SEGUIMIENTO DE PROSPECTOS
# ════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("📋 Seguimiento")
ws5.sheet_view.showGridLines = False
cols5 = [3, 5, 22, 16, 14, 14, 16, 14, 14, 22, 3]
for i, w in enumerate(cols5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.row_dimensions[1].height = 10
ws5.merge_cells("B2:K2")
c = ws5["B2"]
c.value = "SEGUIMIENTO DE PROSPECTOS — SYNERGY BLACK ACCESS"
c.fill = fill(NEGRO)
c.font = Font(bold=True, color=ROJO[2:], size=16, name="Calibri")
c.alignment = center()
ws5.row_dimensions[2].height = 40

ws5.merge_cells("B3:K3")
c = ws5["B3"]
c.value = "Bootcamp 5, 6 y 7 de Junio 2026"
c.fill = fill(GRIS_OSC)
c.font = font(italic=True, size=10)
c.alignment = center()
ws5.row_dimensions[3].height = 20

headers5 = ["#", "NOMBRE", "TELÉFONO / WA", "CANAL ORIGEN",
            "DÍA INTERÉS", "TIPO (S/D)", "ESTADO", "FECHA SEGUIM.",
            "MONTO ($)", "NOTAS"]
row = 4
for i, h in enumerate(headers5, 2):
    style_header(ws5, row, i, h, bg=NEGRO, size=9)
ws5.row_dimensions[row].height = 28
row += 1

estados_color = {
    "🔥 Interesado":   "FFFF6B35",
    "📞 En seguimiento": "FFFFD700",
    "💰 Apartado":     "FF4FC3F7",
    "✅ Cerrado":      "FF66BB6A",
    "❌ Perdido":      "FFEF9A9A",
}

for i in range(1, 101):
    bg = GRIS_CLARO if i % 2 == 0 else BLANCO
    style_value(ws5, row, 2, i, bg=GRIS_OSC, fg=BLANCO, center_align=True)
    for col in range(3, 12):
        c = ws5.cell(row=row, column=col)
        c.fill = fill(bg)
        c.font = font(color=NEGRO[2:], size=9)
        c.alignment = left()
        c.border = border_thin()
        if col == 9:  # Monto
            c.number_format = '$#,##0'
        elif col == 8:  # Fecha
            c.number_format = 'DD/MM/YYYY'
    ws5.row_dimensions[row].height = 18
    row += 1

row += 1
ws5.merge_cells(f"B{row}:K{row}")
c = ws5.cell(row=row, column=2)
c.value = "ESTADOS: 🔥 Interesado  |  📞 En seguimiento  |  💰 Apartado  |  ✅ Cerrado  |  ❌ Perdido"
c.fill = fill("FFFFF8E1")
c.font = Font(italic=True, color="FF795548", size=9, name="Calibri")
c.alignment = center()
ws5.row_dimensions[row].height = 24


# ════════════════════════════════════════════════════════════════════
# HOJA 6 — KPIs RECOMENDADOS
# ════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("💡 KPIs Recomendados")
ws6.sheet_view.showGridLines = False
for col, w in zip("ABCDE", [3, 35, 25, 40, 3]):
    ws6.column_dimensions[get_column_letter(ord(col)-64)].width = w

ws6.row_dimensions[1].height = 10
ws6.merge_cells("B2:D2")
c = ws6["B2"]
c.value = "KPIs RECOMENDADOS — GUÍA DE MÉTRICAS"
c.fill = fill(NEGRO)
c.font = Font(bold=True, color=ROJO[2:], size=16, name="Calibri")
c.alignment = center()
ws6.row_dimensions[2].height = 40

kpis_guide = [
    ("FASE 1: ADQUISICIÓN DE LEADS", ROJO, [
        ("Costo por Lead (CPL)", "Inversión total / Leads generados", "Benchmark: <$15 MXN por lead en Meta Ads"),
        ("Tasa de confirmación", "Confirmados / Leads totales", "Ene 2025: 39.86% — Meta: >40%"),
        ("Leads por canal", "Desglose por fuente de tráfico", "Identifica tu canal más eficiente"),
        ("Costo por confirmado", "Inversión / Confirmados", "Más preciso que CPL para medir calidad"),
    ]),
    ("FASE 2: ACTIVACIÓN (DURANTE EL EVENTO)", NARANJA, [
        ("Tasa de asistencia", "Visualizaciones / Confirmados", "Ene 2025: 15.55% — Meta: >18%"),
        ("Pico de audiencia", "Máximo simultáneo de viewers", "Indica el mejor momento para lanzar la oferta"),
        ("Retención entre días", "Asistentes Día N / Asistentes Día N-1", "Mide el engagement del contenido"),
        ("Interesados capturados", "Leads calientes que piden info del Black", "Base de datos de seguimiento"),
        ("Drop-off por hora", "Salidas durante el evento", "Identifica qué contenido pierde audiencia"),
    ]),
    ("FASE 3: CONVERSIÓN (VENTAS)", NEGRO, [
        ("Tasa de conversión global", "Blacks vendidos / Leads totales", "Ene 2025: 0.36% — Micro pero poderosa"),
        ("Conversión sobre asistentes", "Blacks / Asistentes", "Ene 2025: 0.47% — La más accionable"),
        ("Ticket promedio", "Ingreso total / Transacciones", "Ene 2025: ~$23,489 MXN"),
        ("Mix Simple vs Doble", "% de dobles del total", "El doble tiene mayor ingreso por transacción"),
        ("Tasa de apartados", "Apartados / Total Black", "Ene 2025: 10.5% — Cuántos quedan pendientes"),
        ("Costo por Black vendido", "Inversión publicitaria / Blacks vendidos", "KPI de eficiencia total de la campaña"),
    ]),
    ("FASE 4: POST-EVENTO Y SEGUIMIENTO", GRIS_MED, [
        ("Tasa de recuperación de apartados", "Cobros recuperados / Total apartados", "Meta: >80% de recuperación"),
        ("Ingreso proyectado total", "Neto + Potencial apartados (80%)", "Para forecast financiero"),
        ("ROI de la campaña", "(Ingreso - Inversión) / Inversión", "El KPI más importante para el negocio"),
        ("NPS del evento", "Satisfacción de los asistentes (0-10)", "Impacta en referidos futuros"),
        ("Tiempo promedio de cierre", "Días desde primer contacto hasta venta", "Optimiza el proceso de seguimiento"),
    ]),
]

row = 4
for fase, color, kpis in kpis_guide:
    ws6.merge_cells(f"B{row}:D{row}")
    c = ws6.cell(row=row, column=2)
    c.value = f"▸  {fase}"
    c.fill = fill(color)
    c.font = font(bold=True, size=11)
    c.alignment = left()
    ws6.row_dimensions[row].height = 26
    row += 1

    style_header(ws6, row, 2, "KPI", bg=GRIS_OSC, size=9)
    style_header(ws6, row, 3, "FÓRMULA / DEFINICIÓN", bg=GRIS_OSC, size=9)
    style_header(ws6, row, 4, "REFERENCIA / BENCHMARK", bg=GRIS_OSC, size=9)
    ws6.row_dimensions[row].height = 22
    row += 1

    for i, (kpi, formula, bench) in enumerate(kpis):
        bg = GRIS_CLARO if i % 2 == 0 else BLANCO
        style_value(ws6, row, 2, kpi, bg=bg, bold=True)
        style_value(ws6, row, 3, formula, bg=bg)
        style_value(ws6, row, 4, bench, bg="FFFEF9EF")
        ws6.row_dimensions[row].height = 22
        row += 1
    row += 1


# ════════════════════════════════════════════════════════════════════
# GUARDAR
# ════════════════════════════════════════════════════════════════════
output_path = "/Users/pablomunoz/Bootcamp/Bootcamp_Metricas_Jun2026.xlsx"
wb.save(output_path)
print(f"✅ Archivo guardado: {output_path}")
