import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NEGRO      = "FF1A1A1A"
ROJO       = "FFD63B2F"
NARANJA    = "FFE8640C"
BLANCO     = "FFFFFFFF"
GRIS_OSC   = "FF2D2D2D"
GRIS_MED   = "FF444444"
GRIS_CLARO = "FFF5F5F5"
VERDE      = "FF2ECC71"
AMARILLO   = "FFF1C40F"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=BLANCO, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def border_thin():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(ws, row, col, value, bg=ROJO, fg=BLANCO, size=11, bold=True, span=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=size)
    cell.alignment = center()
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, 26)
    return cell

def style_label(ws, row, col, value, bg=GRIS_OSC, fg=BLANCO, bold=False, span=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=10)
    cell.alignment = left()
    cell.border = border_thin()
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return cell

def style_value(ws, row, col, value, bg=GRIS_CLARO, fg=NEGRO, bold=False, center_align=False, span=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=10)
    cell.alignment = center() if center_align else left()
    cell.border = border_thin()
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return cell

def section_title(ws, row, col, value, bg=ROJO, span=4):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)
    c = ws.cell(row=row, column=col, value=f"▸  {value}")
    c.fill = fill(bg)
    c.font = font(bold=True, size=11)
    c.alignment = left()
    ws.row_dimensions[row].height = 26
    return c


# ── Abrir archivo existente ──────────────────────────────────────────
wb = openpyxl.load_workbook("/Users/pablomunoz/Bootcamp/Bootcamp_Metricas_Jun2026.xlsx")

# Eliminar si ya existe
if "👤 Buyer Persona Black" in wb.sheetnames:
    del wb["👤 Buyer Persona Black"]

ws = wb.create_sheet("👤 Buyer Persona Black")
ws.sheet_view.showGridLines = False

# Anchos de columna
col_widths = [3, 32, 38, 38, 3]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 10

# ── TÍTULO ───────────────────────────────────────────────────────────
ws.merge_cells("B2:D2")
c = ws["B2"]
c.value = "PERFIL DEL COMPRADOR — SYNERGY BLACK ACCESS"
c.fill = fill(NEGRO)
c.font = Font(bold=True, color=ROJO[2:], size=18, name="Calibri")
c.alignment = center()
ws.row_dimensions[2].height = 48

ws.merge_cells("B3:D3")
c = ws["B3"]
c.value = "Análisis basado en el chat de WhatsApp del grupo Black MX 2025 · 210 compradores reales"
c.fill = fill(NEGRO)
c.font = font(italic=True, size=10)
c.alignment = center()
ws.row_dimensions[3].height = 22

ws.row_dimensions[4].height = 10

# ════════════════════════════════════════════════════════════════════
# SECCIÓN 1 — DEMOGRAFÍA
# ════════════════════════════════════════════════════════════════════
row = 5
section_title(ws, row, 2, "DEMOGRAFÍA", bg=ROJO, span=3)
row += 1

style_header(ws, row, 2, "VARIABLE", bg=NEGRO, size=9)
style_header(ws, row, 3, "DATOS", bg=NEGRO, size=9, span=2)
ws.row_dimensions[row].height = 22
row += 1

demo = [
    ("Edad",                   "30 a 55 años — mayoría entre 35 y 48 años"),
    ("Género",                 "Mixto — ligera mayoría mujeres"),
    ("País",                   "México 80%  ·  USA 15%  ·  España / otros 5%"),
    ("Ciudades MX principales","GDL · CDMX · Monterrey · Chihuahua / Cd. Juárez · Puebla · Durango · Cancún · Querétaro · Hidalgo · Oaxaca · Mérida · Aguascalientes"),
    ("Ciudades USA principales","Los Ángeles · San Francisco · Orlando · Miami · California"),
]
for label, val in demo:
    bg = GRIS_CLARO if row % 2 == 0 else BLANCO
    style_label(ws, row, 2, label)
    style_value(ws, row, 3, val, bg=bg, span=2)
    ws.row_dimensions[row].height = 22
    row += 1

row += 1

# ════════════════════════════════════════════════════════════════════
# SECCIÓN 2 — CLUSTERS DE PROFESIÓN
# ════════════════════════════════════════════════════════════════════
section_title(ws, row, 2, "CLUSTERS DE PROFESIÓN", bg=NARANJA, span=3)
row += 1

style_header(ws, row, 2, "CLUSTER", bg=NEGRO, size=9)
style_header(ws, row, 3, "PERFILES IDENTIFICADOS", bg=NEGRO, size=9)
style_header(ws, row, 4, "% EST.", bg=NEGRO, size=9)
ws.row_dimensions[row].height = 22
row += 1

clusters = [
    ("🩺 Salud & Bienestar",        "Médicos, anestesióloga, quiropráctico, psicoterapeuta, enfermera con clínica, terapeuta holística, yoga/pilates, medicina genómica", "25%"),
    ("🏗️ Empresarios tradicionales","Construcción, distribución de alimentos, ferretería, agua, paquetería, ortopedia, florería, cerámica, moda infantil, restaurantes",  "25%"),
    ("📈 Finanzas & Inversiones",    "Traders bolsa (alumnos Cardona), inversión inmobiliaria, asesores patrimoniales, bienes raíces USA",                                  "15%"),
    ("🌐 Network Marketing / MLM",   "Salud y bienestar: IMMUNOTEC, suplementos, tecnología epigenética, multinivel",                                                       "15%"),
    ("💻 Digital & Tecnología",      "Marketing digital, IA y automatización, agencias SEO/Ads, desarrollo de negocios digitales",                                          "10%"),
    ("⚖️ Servicios profesionales",   "Abogados, contadores, consultores ambientales, biólogos, consultores de seguridad e higiene",                                         "5%"),
    ("🎯 Coaches & Formadores",      "Coaches de alto rendimiento, desarrollo humano, formadores, speakers en proceso",                                                      "5%"),
]
for label, perfiles, pct in clusters:
    bg = GRIS_CLARO if row % 2 == 0 else BLANCO
    style_label(ws, row, 2, label, bold=True)
    style_value(ws, row, 3, perfiles, bg=bg)
    style_value(ws, row, 4, pct, bg=bg, center_align=True, bold=True)
    ws.row_dimensions[row].height = 30
    row += 1

row += 1

# ════════════════════════════════════════════════════════════════════
# SECCIÓN 3 — COMPORTAMIENTOS CLAVE
# ════════════════════════════════════════════════════════════════════
section_title(ws, row, 2, "COMPORTAMIENTOS Y MOTIVACIONES CLAVE", bg=NEGRO, span=3)
row += 1

style_header(ws, row, 2, "INSIGHT", bg=GRIS_OSC, size=9)
style_header(ws, row, 3, "DESCRIPCIÓN", bg=GRIS_OSC, size=9)
style_header(ws, row, 4, "IMPLICACIÓN PARA VENTAS", bg=GRIS_OSC, size=9)
ws.row_dimensions[row].height = 22
row += 1

insights = [
    ("1. Ya son compradores de educación",
     "La mayoría toma cursos con Alejandro Cardona (bolsa), Karla Zorola, Titto Gálvez, Brendon Burchard. Invierten en sí mismos de forma recurrente.",
     "El copy NO debe convencerlos de que la educación vale — ya lo creen. Debe mostrar POR QUÉ este evento es el siguiente paso lógico."),
    ("2. Quieren escalar, no empezar",
     "No son principiantes. Tienen negocios funcionando y buscan herramientas para llevarlos al siguiente nivel.",
     "Hablarle a alguien que ya arrancó y siente que está dejando dinero sobre la mesa. Evitar lenguaje de 'cero a héroe'."),
    ("3. El networking vale tanto como el contenido",
     "Desde el día 1 del chat organizaban cenas, grupos de hotel y desayunos. Compran la experiencia social tanto como la formación.",
     "'Las personas que conocerás valen más que cualquier conferencia.' Vender el acceso al ecosistema, no solo a los speakers."),
    ("4. Compran en pareja o con socio",
     "Altísima frecuencia de matrimonios y socios asistiendo juntos al evento.",
     "El Black Doble ($55,497) es su formato natural. Venderlo como decisión inteligente vs. comprar dos simples ($73,994). Ahorro real: $18,497."),
    ("5. Hay recompras — base caliente",
     "Varios miembros tenían la etiqueta 'Synergy Black 2024', indicando que son compradores del año anterior.",
     "Crear campaña específica de retención para Black 2024. Son los más fáciles de cerrar — ya vivieron la experiencia."),
    ("6. Hablan el lenguaje de Sinergéticos",
     "Usan términos como 'ecosistema', 'sinergia', '1+1=3', 'billetarse', 'victoria'. Ya son parte de la cultura.",
     "Usar ese mismo lenguaje en copys. No explicar la filosofía — ellos ya la viven."),
]
for label, desc, impl in insights:
    bg = GRIS_CLARO if row % 2 == 0 else BLANCO
    style_label(ws, row, 2, label, bold=True)
    style_value(ws, row, 3, desc, bg=bg)
    style_value(ws, row, 4, impl, bg="FFFEF9EF")
    ws.row_dimensions[row].height = 52
    row += 1

row += 1

# ════════════════════════════════════════════════════════════════════
# SECCIÓN 4 — DOLORES Y DESEOS
# ════════════════════════════════════════════════════════════════════
section_title(ws, row, 2, "DOLORES Y DESEOS PRINCIPALES", bg=ROJO, span=3)
row += 1

style_header(ws, row, 2, "DOLOR (lo que los frena)", bg=NEGRO, size=9, span=2)
style_header(ws, row, 4, "DESEO (lo que buscan)", bg=NARANJA, size=9)
ws.row_dimensions[row].height = 22
row += 1

dolores = [
    ("No saber usar redes sociales ni marketing digital",    "Aprender a vender más usando internet sin depender de agencias"),
    ("Querer salir del empleo y vivir 100% de su negocio",  "Libertad financiera y de tiempo — ser su propio jefe"),
    ("Querer aprender a invertir (bolsa, bienes raíces)",    "Que su dinero trabaje para ellos — ingresos pasivos"),
    ("Sentir que su negocio está estancado",                 "Escalar a niveles que parecían imposibles"),
    ("Trabajar solos, sin red de apoyo",                     "Rodearse de personas de su nivel o superior"),
    ("Falta de estructura y capital para crecer",            "Acceso a inversores, socios y oportunidades reales"),
]
for dolor, deseo in dolores:
    bg = GRIS_CLARO if row % 2 == 0 else BLANCO
    style_value(ws, row, 2, f"✗  {dolor}", bg=bg, span=2)
    style_value(ws, row, 4, f"✓  {deseo}", bg="FFEBF5EB")
    ws.row_dimensions[row].height = 28
    row += 1

row += 1

# ════════════════════════════════════════════════════════════════════
# SECCIÓN 5 — SEGMENTOS ESPECIALES
# ════════════════════════════════════════════════════════════════════
section_title(ws, row, 2, "SEGMENTOS ESPECIALES A ACTIVAR", bg=NEGRO, span=3)
row += 1

style_header(ws, row, 2, "SEGMENTO", bg=GRIS_OSC, size=9)
style_header(ws, row, 3, "CARACTERÍSTICA", bg=GRIS_OSC, size=9)
style_header(ws, row, 4, "ESTRATEGIA RECOMENDADA", bg=GRIS_OSC, size=9)
ws.row_dimensions[row].height = 22
row += 1

segmentos = [
    ("⭐ Recompras (Black 2024)",
     "Ya vivieron la experiencia — alta intención de volver",
     "Campaña de acceso anticipado con precio especial o beneficio exclusivo para quienes ya fueron Black"),
    ("🇺🇸 Mercado USA",
     "15% del grupo era de USA — hispanos con mayor poder adquisitivo",
     "Enfatizar que es el evento de negocios en español más grande — valor único para la diáspora hispana"),
    ("👫 Parejas y socios",
     "Muchos van en pareja o dúo de socios de negocio",
     "Activar el Black Doble desde el inicio — presentarlo como la opción inteligente, no como upgrade"),
    ("🩺 Profesionales de salud",
     "25% del grupo — médicos, terapeutas, enfermeras con clínica",
     "Copy específico: 'Tienes la vocación, ahora construye el negocio que mereces'"),
    ("📚 Alumnos de Cardona",
     "Alta presencia — ya confían en el ecosistema Sinergéticos",
     "Activar el canal de Cardona como referidor — el alumno que trae más alumnos"),
]
for seg, caract, estrategia in segmentos:
    bg = GRIS_CLARO if row % 2 == 0 else BLANCO
    style_label(ws, row, 2, seg, bold=True)
    style_value(ws, row, 3, caract, bg=bg)
    style_value(ws, row, 4, estrategia, bg="FFFEF9EF")
    ws.row_dimensions[row].height = 36
    row += 1

row += 1

# ════════════════════════════════════════════════════════════════════
# SECCIÓN 6 — FRASES REALES
# ════════════════════════════════════════════════════════════════════
section_title(ws, row, 2, "FRASES REALES DEL CHAT — VOZ DEL CLIENTE", bg=NARANJA, span=3)
row += 1

style_header(ws, row, 2, "FRASE REAL (extraída del grupo WhatsApp)", bg=NEGRO, size=9, span=2)
style_header(ws, row, 4, "QUÉ REVELA", bg=NEGRO, size=9)
ws.row_dimensions[row].height = 22
row += 1

frases = [
    ('"Quiero hacer crecer mi ecosistema"',
     "Buscan red de contactos estratégicos, no solo conocimiento"),
    ('"El año pasado me cambió la vida, este año me propuse estar en el Black"',
     "La experiencia del evento genera recompra — el evento en sí es el mejor argumento de venta"),
    ('"Quiero aprender a llevar mi negocio más lejos e incursionar en internet"',
     "El dolor de estar en el mundo físico y no saber escalar digitalmente es generalizado"),
    ('"No sé de finanzas ni de inversiones, pero me interesa aprender"',
     "Hay un segmento aspiracional que compra por la transformación aunque no tenga negocio grande aún"),
    ('"Soy empleado y quiero renunciar gracias a mis negocios"',
     "La libertad financiera como meta — el Black es el catalizador del salto"),
    ('"Las conexiones que saldrán de aquí nos van a llevar al siguiente nivel"',
     "Se venden el evento entre sí — el networking es el producto real"),
]
for frase, revela in frases:
    style_value(ws, row, 2, frase, bg="FF1A1A2E", fg=AMARILLO[2:], bold=True, span=2)
    style_value(ws, row, 4, revela, bg=GRIS_CLARO)
    ws.row_dimensions[row].height = 36
    row += 1

row += 1

# ════════════════════════════════════════════════════════════════════
# SECCIÓN 7 — RECOMENDACIONES PARA COPYS
# ════════════════════════════════════════════════════════════════════
section_title(ws, row, 2, "RECOMENDACIONES PARA COPYS Y ESTRATEGIA", bg=ROJO, span=3)
row += 1

recomendaciones = [
    ("1. Ángulo principal",
     "Hablarle al empresario/profesional que ya tiene algo funcionando pero siente que el techo está cerca. NO al principiante.",
     "COPY: '¿Ya tienes negocio pero sientes que podrías ir más rápido? Esto es para ti.'"),
    ("2. Prueba social",
     "Usar testimonios de compradores Black reales — especialmente médicos, empresarios y personas que volvieron por segunda vez.",
     "COPY: 'Karla, médico de Durango, va por su tercer año en Black — ¿qué sabe ella que tú aún no?'"),
    ("3. Black Doble como ancla",
     "Presentar el Doble primero — hace que el Simple parezca más accesible. Ancla de precio hacia abajo.",
     "COPY: 'Ven con tu pareja o socio y ahorren $18,497 — porque los grandes negocios se construyen en equipo.'"),
    ("4. Networking como promesa central",
     "No solo los speakers. La cena de gala, los desayunos y las conexiones son parte del producto.",
     "COPY: 'El contacto que harás en la cena de gala puede cambiar tu negocio para siempre.'"),
    ("5. Urgencia real",
     "Los compradores organizan vuelos, hotel y logística con anticipación — la escasez de lugares Black es creíble.",
     "COPY: 'Los primeros 50 Black ya eligieron hotel. ¿Cuándo reservas el tuyo?'"),
    ("6. Lenguaje a usar",
     "'Ecosistema' · 'Sinergia' · 'Siguiente nivel' · 'Escalar' · '1+1=3' · 'Capital social' · 'Victoria'",
     "Evitar: 'aprende a ganar dinero', 'hazte rico', 'curso' — usan lenguaje de comunidad, no de curso."),
    ("7. Mercado USA",
     "15% del grupo era de USA — tienen mayor poder adquisitivo y menos opciones en español.",
     "COPY: 'El evento de negocios más poderoso en español — donde la comunidad hispana juega en grande.'"),
]
for num, (titulo, desc, ejemplo) in enumerate(recomendaciones):
    bg = GRIS_CLARO if num % 2 == 0 else BLANCO
    style_label(ws, row, 2, titulo, bold=True, bg=GRIS_OSC)
    style_value(ws, row, 3, desc, bg=bg)
    style_value(ws, row, 4, ejemplo, bg="FFFFF8E1", fg="FF5D4037")
    ws.row_dimensions[row].height = 44
    row += 1

# Nota al pie
row += 1
ws.merge_cells(f"B{row}:D{row}")
c = ws.cell(row=row, column=2)
c.value = "💡 Análisis generado a partir del chat de WhatsApp del grupo SUMX BLACK ACCESS 2025 · Julio-Agosto 2025 · 210 compradores reales"
c.fill = fill("FFFFF8E1")
c.font = Font(italic=True, color="FF795548", size=9, name="Calibri")
c.alignment = left()
ws.row_dimensions[row].height = 24

# ── Guardar ──────────────────────────────────────────────────────────
wb.save("/Users/pablomunoz/Bootcamp/Bootcamp_Metricas_Jun2026.xlsx")
print("✅ Hoja 'Buyer Persona Black' agregada al Excel.")
